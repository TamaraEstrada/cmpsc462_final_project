import matplotlib.pyplot as plt
import networkx as nx
from openpyxl import load_workbook

workbook = load_workbook(filename="Available_Classes.xlsx")
sheet = workbook.active
available_classes = dict()
for row in sheet.iter_rows(min_row=2, values_only=True):
    course = row[0]
    course_info = {
        "credits": row[1],
        "credit_type": row[2],
        "semesters": list(row[3].split(", ")),
        # list comprehension to include only non-None values
        "prerequisites": [prereq for prereq in row[4:] if prereq is not None]
        #  "prerequisites": row[4:]
    }
    # print(course, credits, credit_type, sep = '\t')
    available_classes[course] = course_info

# print(available_classes)
for course, info in available_classes.items():
    print(course, info, sep=": ")
    print()  # This will print a new line after each entry


def create_course_graph(courses):
    # creates and returns a directed graph from the given course information.
    G = nx.DiGraph()
    # Add nodes and edges to the graph
    for course, info in courses.items():
        G.add_node(course)
        for prereq in info["prerequisites"]:
            G.add_edge(prereq, course)  # Edge from prerequisite to the course

    return G


# Create the graph
G = create_course_graph(available_classes)
# Draw the graph
nx.draw(G, with_labels=True)
plt.show()


def schedule_course(available, taken, current, sheet):
    course = input("What course would you like to schedule? ")
    print(course)
    if course in available:
        if course not in taken:
            semester = input("Fall or Spring semester? ")
            if semester in available[course]["semesters"]:
                current_credits = available[course]["credits"]
                for c in current:
                    current_credits += available[c]["credits"]
                print(current_credits)
                if current_credits < 19:
                    valid = True
                    for prereq in available[course]["prerequisites"]:
                        if prereq not in taken and prereq is not None:
                            valid = False
                            print(f"{prereq} has not been taken/scheduled yet.")
                    if valid:
                        credits = available[course]["credits"]
                        type = available[course]["credit_type"]
                        print(course, credits, type, sep="\t")
                        sheet.insert_rows(idx=1)
                        sheet["A1"] = course
                        sheet["B1"] = credits
                        sheet["C1"] = type
                        workbook.save("TestSchedule.xlsx")
                        taken[course] = available[course]
                        current[course] = available[course]
                        print("Course scheduled!")
                else:
                    print("Course would exceed credit limit for semester.")
            else:
                print("Course unavailable in given semester.")
        else:
            print("Course has already been scheduled/taken.")
    else:
        print("Invalid course name.")
    # print(course, credits, type, sep = '\t')


workbook = load_workbook(filename="TestSchedule.xlsx")
sheet = workbook.active
taken_classes = dict()
for row in sheet.iter_rows(values_only=True):
    course = row[0]
    course_info = {"credits": row[1], "credit_type": row[2]}
    taken_classes[course] = course_info

current_classes = dict()
while 1:
    x = input("Would you like to schedule a class? ")
    if x == "yes":
        schedule_course(available_classes, taken_classes, current_classes, sheet)
    else:
        print("Okay.")
        break
